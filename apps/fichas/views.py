
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .models import FichaEvaluacion, FamiliarDelEvaluado, FichaDetalle, Pregunta, Opcion, Dimension, Institucion, FichaHistorial
from apps.ubigeo.models import Departamento
from datetime import timedelta
from django.utils import timezone
from django.db.models import Q
from .forms import PreguntaForm, OpcionFormSet
from django.db import transaction
#apartado del administrador
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q

# Importar modelos y forms
from .models import Institucion, Dimension, Pregunta
from .forms import InstitucionForm, DimensionForm, PreguntaForm

from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

def es_admin(user):
    return user.is_authenticated and user.rol == 'ADMIN'


# =======================================================

@login_required 
def mis_encuestas(request):
    """
    Dashboard con Búsqueda y Filtros.
    """
    # 1. BASE: Fichas del usuario
    mis_fichas = FichaEvaluacion.objects.filter(
        usuario_registra=request.user
    ).order_by('-fecha_registro')

    # 2. CAPTURAR PARÁMETROS DE BÚSQUEDA
    dni_query = request.GET.get('dni', '').strip()
    fecha_filtro = request.GET.get('rango_fecha', '')
    
    # 3. APLICAR FILTROS (Si existen)
    
    # Filtro por DNI (Búsqueda exacta o parcial)
    if dni_query:
        mis_fichas = mis_fichas.filter(dni_evaluado__icontains=dni_query)

    # Filtro por Fecha
    hoy = timezone.now().date()
    if fecha_filtro == 'hoy':
        mis_fichas = mis_fichas.filter(fecha_registro__date=hoy)
    elif fecha_filtro == '7dias':
        semana_atras = hoy - timedelta(days=7)
        mis_fichas = mis_fichas.filter(fecha_registro__date__gte=semana_atras)
    elif fecha_filtro == 'mes':
        mes_atras = hoy - timedelta(days=30)
        mis_fichas = mis_fichas.filter(fecha_registro__date__gte=mes_atras)

    # 4. CALCULAR KPI's (Sobre el total histórico, NO sobre la búsqueda, para no perder contexto)
    # Nota: Si prefieres que los KPIs cambien según la búsqueda, usa 'mis_fichas' en lugar de 'todas_mis_fichas'
    todas_mis_fichas = FichaEvaluacion.objects.filter(usuario_registra=request.user)
    
    total = todas_mis_fichas.count()
    bajo = todas_mis_fichas.filter(nivel_riesgo='RIESGO BAJO').count()
    moderado = todas_mis_fichas.filter(nivel_riesgo='RIESGO MODERADO').count()
    severo = todas_mis_fichas.filter(nivel_riesgo='RIESGO SEVERO').count()
    critico = todas_mis_fichas.filter(nivel_riesgo='RIESGO CRÍTICO').count()
   
# if puntaje_total >= 126: ficha.nivel_riesgo = 'RIESGO CRÍTICO'
#                 elif puntaje_total >= 76: ficha.nivel_riesgo = 'RIESGO SEVERO'
#                 elif puntaje_total >= 26: ficha.nivel_riesgo = 'RIESGO MODERADO'
#                 else: ficha.nivel_riesgo = 'RIESGO BAJO'
    context = {
        'fichas': mis_fichas,
        'total': total,
        'bajo': bajo,
        'moderado': moderado,
        'severo': severo,
        'critico': critico, 
        # Devolvemos los valores para mantenerlos en los inputs tras buscar
        'filtro_dni': dni_query,
        'filtro_fecha': fecha_filtro
    }
    
    return render(request, 'fichas/dashboard_encuestador.html', context)
# =======================================================
# PARTE 2: MODELO PRINCIPAL DE LA FICHA


@login_required
def registrar_ficha(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. CREAR FICHA
                ficha = FichaEvaluacion.objects.create(
                    usuario_registra=request.user,
                    institucion=Institucion.objects.first(),
                    
                    # Datos Personales
                    nombres_evaluado=request.POST.get('nombres_encuestado', '').split(' ')[0], 
                    apellidos_evaluado=' '.join(request.POST.get('apellidos_encuestado', '').split(' ')[1:]),
                    edad_evaluado=request.POST.get('edad') or 0,
                    fecha_nacimiento=request.POST.get('fecha_nacimiento'),
                    dni_evaluado=request.POST.get('dni'),
                    sexo_evaluado=request.POST.get('sexo'),
                    nivel_educativo=request.POST.get('nivel_educativo'),
                    
                    
                    # UBICACIÓN ACTUAL (Solo guardamos esto)
                    direccion_domicilio=request.POST.get('direccion'),
                    ubigeo_departamento_id=request.POST.get('ubigeo_departamento') or None,
                    ubigeo_provincia_id=request.POST.get('ubigeo_provincia') or None,
                    ubigeo_distrito_id=request.POST.get('ubigeo_distrito') or None,
                    
                    # Contacto
                    telefono_contacto=request.POST.get('telefono'),
                    email_contacto=request.POST.get('email'),
                    emergencia_nombres=request.POST.get('emergencia_contacto'),
                    emergencia_telefono=request.POST.get('emergencia_telefono'),
                    emergencia_parentesco=request.POST.get('emergencia_parentesco'),
                    
                    # Familia
                    jefe_hogar=request.POST.get('jefe_hogar'),
                    num_integrantes=request.POST.get('num_integrantes') or 0,
                    observaciones_familia=request.POST.get('observaciones_familia'),
                )

                # 2. FAMILIARES
                total_filas = int(request.POST.get('total_familiares', 0))
                for i in range(1, total_filas + 1):
                    nombre = request.POST.get(f'fam_{i}_nombre')
                    if nombre:
                        FamiliarDelEvaluado.objects.create(
                            ficha=ficha,
                            nombres=nombre,
                            parentesco=request.POST.get(f'fam_{i}_parentesco', ''),
                            edad=request.POST.get(f'fam_{i}_edad') or 0,
                            sexo=request.POST.get(f'fam_{i}_sexo', 'M'),
                            estado_civil=request.POST.get(f'fam_{i}_ecivil', ''),
                            nivel_educativo=request.POST.get(f'fam_{i}_neducativo', ''),
                            ocupacion=request.POST.get(f'fam_{i}_ocupacion', ''),
                            ingresos=request.POST.get(f'fam_{i}_ingresos') or 0
                        )

                # 3. PREGUNTAS
                preguntas = Pregunta.objects.all()
                puntaje_total = 0
                for pregunta in preguntas:
                    opcion_id = request.POST.get(f'pregunta_{pregunta.id}')
                    if opcion_id:
                        opcion = Opcion.objects.get(id=opcion_id)
                        FichaDetalle.objects.create(
                            ficha=ficha,
                            pregunta=pregunta,
                            opcion_seleccionada=opcion,
                            puntaje_obtenido=opcion.puntaje
                        )
                        puntaje_total += opcion.puntaje

                # 4. CALCULAR RIESGO
                ficha.puntaje_total = puntaje_total
                if puntaje_total >= 126: ficha.nivel_riesgo = 'RIESGO CRÍTICO'
                elif puntaje_total >= 76: ficha.nivel_riesgo = 'RIESGO SEVERO'
                elif puntaje_total >= 26: ficha.nivel_riesgo = 'RIESGO MODERADO'
                else: ficha.nivel_riesgo = 'RIESGO BAJO'
                
                ficha.save()
                messages.success(request, f'Ficha guardada. Riesgo: {ficha.nivel_riesgo}')
                return redirect('mis_encuestas')

        except Exception as e:
            messages.error(request, f'Error al guardar: {str(e)}')
            dimensiones = Dimension.objects.prefetch_related('preguntas__opciones').order_by('orden')
            departamentos = Departamento.objects.all().order_by('nombre')
            return render(request, 'fichas/form_riesgo.html', {'dimensiones': dimensiones, 'departamentos': departamentos})

    # GET
    dimensiones = Dimension.objects.prefetch_related('preguntas__opciones').order_by('orden')
    departamentos = Departamento.objects.all().order_by('nombre')
    return render(request, 'fichas/form_riesgo.html', {'dimensiones': dimensiones, 'departamentos': departamentos})

from django.shortcuts import render, redirect, get_object_or_404
# ... tus otros imports ...

@login_required
def ver_ficha(request, ficha_id):
    # 1. Obtener la ficha (o 404 si no existe)
    ficha = get_object_or_404(FichaEvaluacion, id=ficha_id)
    
    # OPCIONAL: Seguridad (Solo ver tus propias fichas)
    # if ficha.usuario_registra != request.user:
    #     return redirect('mis_encuestas')

    # 2. Obtener familiares
    familiares = ficha.familiares.all()

    ficha_historial = FichaHistorial.objects.filter(ficha=ficha).order_by('-fecha_edicion')
    respuestas = FichaDetalle.objects.filter(ficha=ficha).select_related(
        'pregunta', 
        'pregunta__dimension', 
        'opcion_seleccionada'
    ).order_by('pregunta__dimension__orden', 'pregunta__orden')

    return render(request, 'fichas/ver_ficha.html', {
        'ficha': ficha,
        'familiares': familiares,
        'respuestas': respuestas,
        'ficha_historial': ficha_historial
    })

from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import FichaEvaluacion

from .models import FichaEvaluacion, Institucion # Importamos Institucion

@login_required
def listar_mis_fichas(request):
    # 1. Obtener parámetros de búsqueda
    # Usamos el ID de la institución para una búsqueda más precisa
    inst_id = request.GET.get('institucion_id', '') 
    search_dni = request.GET.get('dni', '')

    # 2. Queryset base
    queryset = FichaEvaluacion.objects.all().select_related('institucion').order_by('-fecha_registro')

    # 3. Aplicar filtros
    if inst_id:
        queryset = queryset.filter(institucion_id=inst_id)
    
    if search_dni:
        queryset = queryset.filter(dni_evaluado__icontains=search_dni)

    # 4. Traer todas las instituciones para el combo
    instituciones = Institucion.objects.all().order_by('nombre')

    # 5. Paginación
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'fichas/listar_fichas.html', {
        'fichas': page_obj,
        'instituciones': instituciones, # Enviamos la lista de instituciones
        'inst_id': inst_id,             # Enviamos el ID seleccionado para mantenerlo en el combo
        'search_dni': search_dni,
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import FichaEvaluacion, FichaHistorial
from .forms import FichaEvaluacionForm

@login_required
def editar_ficha(request, ficha_id):
    # 1. Traemos la ficha original de la DB
    ficha = get_object_or_404(FichaEvaluacion, id=ficha_id)
    dimensiones = Dimension.objects.prefetch_related('preguntas__opciones').all().order_by('orden')
    
    # 2. Diccionario de respuestas actuales para auditar preguntas
    detalles_actuales = {
        d.pregunta_id: d.opcion_seleccionada.texto 
        for d in ficha.detalles.select_related('opcion_seleccionada').all()
    }

    if request.method == 'POST':
        form = FichaEvaluacionForm(request.POST, instance=ficha)
        
        if form.is_valid():
            cambios_detectados = []

            # --- CORRECCIÓN DE AUDITORÍA DE CABECERA ---
            # Obtenemos una copia "congelada" de la ficha tal como está en la DB
            ficha_original = FichaEvaluacion.objects.get(pk=ficha.pk)

            for campo in form.changed_data:
                # Sacamos el valor de la copia que NO ha sido tocada por el formulario
                valor_antiguo = getattr(ficha_original, campo)
                valor_nuevo = form.cleaned_data.get(campo)
                
                # Solo registramos si realmente son diferentes (evita ruidos)
                if valor_antiguo != valor_nuevo:
                    cambios_detectados.append(f"Campo {campo}: '{valor_antiguo}' ➔ '{valor_nuevo}'")

            # Ahora sí, preparamos el guardado
            ficha = form.save(commit=False)
            nuevo_puntaje_total = 0

            # --- AUDITORÍA DE PREGUNTAS (MINUCIOSA) ---
            for key, value in request.POST.items():
                if key.startswith('pregunta_'):
                    id_pregunta = int(key.split('_')[1])
                    id_opcion_nueva = int(value)
                    
                    pregunta_obj = Pregunta.objects.get(id=id_pregunta)
                    opcion_nueva_obj = Opcion.objects.get(id=id_opcion_nueva)
                    
                    texto_anterior = detalles_actuales.get(id_pregunta, "Sin respuesta")
                    texto_nuevo = opcion_nueva_obj.texto

                    if texto_anterior != texto_nuevo:
                        cambios_detectados.append(
                            f"Pregunta {pregunta_obj.orden}: '{texto_anterior}' ➔ '{texto_nuevo}'"
                        )
                    
                    FichaDetalle.objects.update_or_create(
                        ficha=ficha, pregunta=pregunta_obj,
                        defaults={
                            'opcion_seleccionada': opcion_nueva_obj,
                            'puntaje_obtenido': opcion_nueva_obj.puntaje
                        }
                    )
                    nuevo_puntaje_total += opcion_nueva_obj.puntaje
            
            ficha.puntaje_total = nuevo_puntaje_total
            ficha.save()

            # 3. Guardar Auditoría solo si hubo cambios reales
            if cambios_detectados:
                FichaHistorial.objects.create(
                    ficha=ficha,
                    usuario=request.user,
                    accion="Edición de Ficha",
                    detalles="\n".join(cambios_detectados)
                )
            
            return redirect('listar_fichas')
    else:
        form = FichaEvaluacionForm(instance=ficha)
    respuestas_ids = list(ficha.detalles.values_list('opcion_seleccionada_id', flat=True))
    return render(request, 'fichas/editar_ficha.html', {
        'form': form,
        'ficha': ficha,
        'dimensiones': dimensiones,
        'respuestas_ids': respuestas_ids,
        # 'respuestas_actuales': {k: v for k, v in ficha.detalles.values_list('pregunta_id', 'opcion_seleccionada_id')}
    })




@login_required
def exportar_excel(request):
    # 1. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Filtrado"

    # 2. Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 3. Encabezados
    headers = ["ID", "Fecha", "Estudiante", "DNI", "Institución", "Puntaje", "Nivel de Riesgo"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # =======================================================
    # 4. OBTENER DATOS Y APLICAR FILTROS (IGUAL QUE EN DASHBOARD)
    # =======================================================
    
    # Base: Fichas del usuario
    fichas = FichaEvaluacion.objects.filter(usuario_registra=request.user).order_by('-fecha_registro')
  
    # Capturar parámetros de la URL
    dni_query = request.GET.get('dni', '').strip()
    fecha_filtro = request.GET.get('rango_fecha', '')

    # Filtro por DNI
    if dni_query:
        fichas = fichas.filter(dni_evaluado__icontains=dni_query)

    # Filtro por Fecha
    hoy = timezone.now().date()
    if fecha_filtro == 'hoy':
        fichas = fichas.filter(fecha_registro__date=hoy)
    elif fecha_filtro == '7dias':
        semana_atras = hoy - timedelta(days=7)
        fichas = fichas.filter(fecha_registro__date__gte=semana_atras)
    elif fecha_filtro == 'mes':
        mes_atras = hoy - timedelta(days=30)
        fichas = fichas.filter(fecha_registro__date__gte=mes_atras)

    # =======================================================

    # 5. Escribir filas con los datos ya filtrados
    for ficha in fichas:
        row = [
            ficha.id,
            ficha.fecha_registro.strftime("%d/%m/%Y %H:%M"),
            f"{ficha.nombres_evaluado} {ficha.apellidos_evaluado}",
            ficha.dni_evaluado,
            ficha.institucion.nombre if ficha.institucion else "-",
            ficha.puntaje_total,
            ficha.nivel_riesgo
        ]
        ws.append(row)

    # 6. Ajustar ancho
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # 7. Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Fichas.xlsx"'
    
    wb.save(response)
    return response




# =======================================================
# 1. GESTIÓN DE INSTITUCIONES
# =======================================================

@user_passes_test(es_admin)
def lista_instituciones(request):
    query = request.GET.get('q', '')
    items = Institucion.objects.all().order_by('-fecha_registro')
    
    if query:
        items = items.filter(Q(nombre__icontains=query) | Q(codigo_modular__icontains=query))
    
    paginator = Paginator(items, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'configuracion/lista_instituciones.html', {
        'items': page_obj, 'query': query
    })

@user_passes_test(es_admin)
def gestion_institucion(request, pk=None):
    # HÍBRIDO: Sirve para CREAR (pk=None) y EDITAR (pk=integro)
    instancia = get_object_or_404(Institucion, pk=pk) if pk else None
    titulo = "Editar Institución" if pk else "Nueva Institución"
    
    if request.method == 'POST':
        form = InstitucionForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, f'Institución {"actualizada" if pk else "creada"} correctamente.')
            return redirect('lista_instituciones')
        else:
            messages.error(request, 'Verifique los errores del formulario.')
    else:
        form = InstitucionForm(instance=instancia)
        
    return render(request, 'configuracion/form_generico.html', {
        'form': form, 'titulo': titulo, 'back_url': 'lista_instituciones'
    })

@user_passes_test(es_admin)
def eliminar_institucion(request, pk):
    item = get_object_or_404(Institucion, pk=pk)
    try:
        item.delete()
        messages.success(request, 'Institución eliminada.')
    except:
        messages.error(request, 'No se puede eliminar porque tiene fichas asociadas.')
    return redirect('lista_instituciones')

# =======================================================
# 2. GESTIÓN DE DIMENSIONES Y PREGUNTAS (Banco de Preguntas)
# =======================================================

@user_passes_test(es_admin)
def lista_banco_preguntas(request):
    """Vista maestra para ver Dimensiones y sus Preguntas"""
    dimensiones = Dimension.objects.all().order_by('orden').prefetch_related('preguntas')
    return render(request, 'configuracion/lista_banco.html', {'dimensiones': dimensiones})

# --- CRUD DIMENSION ---
@user_passes_test(es_admin)
def gestion_dimension(request, pk=None):
    instancia = get_object_or_404(Dimension, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = DimensionForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dimensión guardada.')
            return redirect('lista_banco_preguntas')
    else:
        form = DimensionForm(instance=instancia)
        
    return render(request, 'configuracion/form_generico.html', {
        'form': form, 'titulo': 'Gestión de Dimensión', 'back_url': 'lista_banco_preguntas'
    })

# --- CRUD PREGUNTA ---
from django.db import transaction
from django.db.models import F, Max 

@user_passes_test(es_admin)
def gestion_pregunta(request, pk=None):
    pregunta = get_object_or_404(Pregunta, pk=pk) if pk else None
    
    # Capturamos datos ORIGINALES para comparar después
    orden_original = pregunta.orden if pregunta else None
    dimension_original_id = pregunta.dimension.id if pregunta else None
    
    if request.method == 'POST':
        form = PreguntaForm(request.POST, instance=pregunta)
        formset = OpcionFormSet(request.POST, instance=pregunta)
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                nueva_pregunta = form.save(commit=False)
                
                nuevo_orden = nueva_pregunta.orden
                nueva_dimension = nueva_pregunta.dimension
                
                # === LÓGICA MAESTRA DE REORDENAMIENTO ===

                # CASO 1: CREACIÓN NUEVA (Insertar)
                if not pk:
                    Pregunta.objects.filter(
                        dimension=nueva_dimension, 
                        orden__gte=nuevo_orden
                    ).update(orden=F('orden') + 1)

                # CASO 2: EDICIÓN
                else:
                    # SUB-CASO 2A: CAMBIO DE DIMENSIÓN (El más complejo)
                    if dimension_original_id != nueva_dimension.id:
                        # 1. Cerrar hueco en la dimensión VIEJA (Restar 1 a los siguientes)
                        Pregunta.objects.filter(
                            dimension_id=dimension_original_id,
                            orden__gt=orden_original
                        ).update(orden=F('orden') - 1)

                        # 2. Abrir espacio en la dimensión NUEVA (Sumar 1 en el destino)
                        Pregunta.objects.filter(
                            dimension=nueva_dimension,
                            orden__gte=nuevo_orden
                        ).update(orden=F('orden') + 1)

                    # SUB-CASO 2B: MISMA DIMENSIÓN, SOLO CAMBIÓ ORDEN
                    elif orden_original != nuevo_orden:
                        if orden_original > nuevo_orden:
                            # Mover arriba: Empujar intermedios abajo
                            Pregunta.objects.filter(
                                dimension=nueva_dimension,
                                orden__gte=nuevo_orden,
                                orden__lt=orden_original
                            ).update(orden=F('orden') + 1)
                        else:
                            # Mover abajo: Subir intermedios arriba
                            Pregunta.objects.filter(
                                dimension=nueva_dimension,
                                orden__gt=orden_original,
                                orden__lte=nuevo_orden
                            ).update(orden=F('orden') - 1)

                # Guardamos finalmente
                nueva_pregunta.save()
                formset.instance = nueva_pregunta
                formset.save()
                
            messages.success(request, f'Pregunta guardada. Posición {nueva_pregunta.orden} en {nueva_pregunta.dimension}.')
            return redirect('lista_banco_preguntas')
    else:
        form = PreguntaForm(instance=pregunta)
        formset = OpcionFormSet(instance=pregunta)

    # Datos para el JS
    ordenes_por_dimension = {}
    for dim in Dimension.objects.all():
        max_orden = Pregunta.objects.filter(dimension=dim).aggregate(Max('orden'))['orden__max']
        ordenes_por_dimension[dim.id] = max_orden if max_orden else 0

    return render(request, 'configuracion/form_pregunta_custom.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Editar Pregunta' if pk else 'Nueva Pregunta',
        'back_url': 'lista_banco_preguntas',
        'ordenes_json': ordenes_por_dimension
    })

@user_passes_test(es_admin)
def eliminar_generico(request, modelo, pk):
    # Función dinámica para borrar Dimensiones o Preguntas
    clases = {'dimension': Dimension, 'pregunta': Pregunta}
    model_class = clases.get(modelo)
    
    if model_class:
        with transaction.atomic():
            item = get_object_or_404(model_class, pk=pk)
            
            # LÓGICA ESPECIAL PARA PREGUNTAS (Cerrar el hueco)
            if modelo == 'pregunta':
                orden_borrado = item.orden
                dim_id = item.dimension.id
                
                # 1. Eliminamos el item
                item.delete()
                
                # 2. Reordenamos: A todos los que estaban DESPUÉS, les restamos 1
                Pregunta.objects.filter(
                    dimension_id=dim_id,
                    orden__gt=orden_borrado
                ).update(orden=F('orden') - 1)
                
                msg = 'Pregunta eliminada y numeración actualizada.'
            
            else:
                # Borrado normal para dimensiones u otros
                item.delete()
                msg = 'Elemento eliminado correctamente.'

            messages.success(request, msg)
    
    return redirect('lista_banco_preguntas')



