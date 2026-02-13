from django.shortcuts import render, redirect, get_list_or_404
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView

# --- IMPORTACIONES FALTANTES ---

from django.contrib.auth.decorators import login_required
from django.contrib import messages

# -------------------------------
from .forms import LoginForm, PerfilForm, UsuarioAdminForm, EncuestadorForm
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from .models import Usuario

from django.core.paginator import Paginator
from django.utils import timezone
from apps.fichas.models import FichaEvaluacion
from datetime import timedelta
# ... (tus otras vistas) ...

from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import TruncDate
from apps.usuarios.models import Usuario
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.shortcuts import redirect


@login_required
def home(request):
    """Punto de entrada único que deriva según el rol."""
    user = request.user
    
    if user.rol == 'ENCUESTADOR':
        return redirect('mis_encuestas')
    
    elif user.rol == 'SUPERVISOR':
        return redirect('dashboard_admin')
    
    elif user.rol == 'ADMIN':
        return redirect('lista_usuarios') # O 'lista_usuarios' según prefieras
    
    # Si por alguna razón no tiene rol, lo mandamos al login o una página neutral
    return redirect('login')

class CustomLoginView(LoginView):
    template_name = 'usuarios/login.html'
    authentication_form = LoginForm
    redirect_authenticated_user = True 



def es_admin(user):
    return user.is_authenticated and user.rol == 'ADMIN' # Ajusta 'ADMIN' a como lo tengas en tu modelo

def es_supervisor(user):
    return user.is_authenticated and user.rol == 'SUPERVISOR'



def es_admin(user):
    return user.is_authenticated and user.rol == 'ADMIN'


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Q, Sum
from django.utils import timezone
from datetime import timedelta
from apps.fichas.models import FichaEvaluacion, Institucion, Pregunta, Dimension
from django.contrib.auth import get_user_model

from django.db.models import Max, Min


User = get_user_model()

@user_passes_test(es_supervisor)
def dashboard_admin(request):
    hoy = timezone.now().date()
  
    
    # --- KPI CARDS (Métricas Principales) ---
    kpi_hoy = 0
    kpi_total_fichas = 0
    kpi_casos_criticos = 0
    kpi_puntaje_mas_bajo = 0
    kpi_puntaje_promedio = 0
    kpi_puntaje_mas_alto = 0
    todas_las_fichas = FichaEvaluacion.objects.filter(usuario_registra__supervisor_asignado=request.user).all()

    
    for ficha in todas_las_fichas:
        kpi_total_fichas += 1
        if ficha.fecha_registro.date() == hoy:
            kpi_hoy += 1
        if ficha.nivel_riesgo == 'RIESGO CRÍTICO' or ficha.nivel_riesgo == 'RIESGO SEVERO':
            kpi_casos_criticos += 1
    kpi_puntaje_mas_bajo = todas_las_fichas.aggregate(minimo=Min('puntaje_total'))['minimo'] or 0
    kpi_puntaje_mas_alto = todas_las_fichas.aggregate(maximo=Max('puntaje_total'))['maximo'] or 0
    kpi_puntaje_promedio = todas_las_fichas.aggregate(promedio=Avg('puntaje_total'))['promedio'] or 0
    
    
    kpi_encuestadores = User.objects.filter(supervisor_asignado=request.user).count()

    # 1. Preparación de variables
    labels_tendencia = []
    data_tendencia = []
    for i in range(7, -1, -1):
        dia = hoy - timedelta(days=i)
        labels_tendencia.append(dia.strftime('%d %b'))

        conteo_dia = 0

        for ficha in todas_las_fichas:
            if ficha.fecha_registro.date() == dia:
                conteo_dia += 1
        data_tendencia.append(conteo_dia)

    # 2. Distribución de Riesgos (Pie Chart)
    riesgos_query = FichaEvaluacion.objects.filter(
    usuario_registra__supervisor_asignado=request.user
    ).values('nivel_riesgo').annotate(total=Count('id'))

    # Generamos las etiquetas y los datos para el gráfico
    riesgos_labels = [r['nivel_riesgo'] or 'Sin Evaluar' for r in riesgos_query]
    riesgos_data = [r['total'] for r in riesgos_query]

    # 3. Top 5 Agentes (Filtrado por el equipo del supervisor logueado)
    top_encuestadores = Usuario.objects.filter(
        supervisor_asignado=request.user # Solo encuestadores que reportan a él
    ).annotate(
        # Total histórico de fichas de cada encuestador del equipo
        total_fichas=Count('fichas_realizadas'), 
        
        # Fichas realizadas por el encuestador solo en el mes actual
        fichas_mes=Count(
            'fichas_realizadas', 
            filter=Q(fichas_realizadas__fecha_registro__month=hoy.month)
        )
    ).order_by('-total_fichas')[:5]

    # 4. Top 5 Instituciones (Filtrado por el equipo del supervisor logueado)
    top_instituciones = Institucion.objects.filter(
        # Filtramos para considerar solo instituciones que tengan fichas de su equipo
        fichaevaluacion__usuario_registra__supervisor_asignado=request.user
    ).annotate(
        # Total de fichas del equipo en esa institución
        total=Count(
            'fichaevaluacion', 
            filter=Q(fichaevaluacion__usuario_registra__supervisor_asignado=request.user)
        )
    ).distinct().order_by('-total')[:5]

    # --- ACTIVIDAD RECIENTE (Últimas 8 fichas del equipo del supervisor) ---
    ultimas_fichas = FichaEvaluacion.objects.filter(
        usuario_registra__supervisor_asignado=request.user
    ).select_related(
        'institucion', 
        'usuario_registra'
    ).order_by('-fecha_registro')[:8]

    context = {
        'kpi_hoy': kpi_hoy,
        'kpi_total_fichas': kpi_total_fichas,
        'kpi_casos_criticos': kpi_casos_criticos,
        'kpi_encuestadores': kpi_encuestadores,
        'kpi_puntaje_mas_bajo': kpi_puntaje_mas_bajo,
        'kpi_puntaje_mas_alto': kpi_puntaje_mas_alto,
        'kpi_puntaje_promedio': round(kpi_puntaje_promedio, 1),
        'ultimas_fichas': ultimas_fichas,
        'top_encuestadores': top_encuestadores,
        'top_instituciones': top_instituciones,
        # Listas para JS
        'labels_tendencia': labels_tendencia,
        'data_tendencia': data_tendencia,
        'riesgos_labels': riesgos_labels,
        'riesgos_data': riesgos_data,
    }
    return render(request, 'usuarios/dashboard.html', context)

@login_required
def editar_perfil(request):
    if request.method == 'POST':
        # 1. Recibimos los datos del usuario
        form = PerfilForm(request.POST, instance=request.user)
        
        if form.is_valid():
            # CASO ÉXITO: Todo bien, guardamos
            form.save()
            messages.success(request, '¡Tus datos se actualizaron correctamente!')
            return redirect('dashboard')
        else:
            # CASO ERROR: Falló la validación (teléfono mal, dni mal, etc.)
            if hasattr(form, 'estilar_errores'):
                form.estilar_errores()
            
            # B. Mandamos mensaje global
            messages.error(request, 'Por favor corrige los errores marcados en rojo.')
    else:
        # GET: El usuario recién entra, cargamos sus datos limpios de la BD
        form = PerfilForm(instance=request.user)

    # Aquí llega el formulario (ya sea limpio del GET o con errores del POST)
    return render(request, 'usuarios/editar_perfil.html', {'form': form})


# Función auxiliar para saber si es admin

# 1. LISTAR USUARIOS
@user_passes_test(es_admin)
def lista_usuarios(request):
    query = request.GET.get('q', '').strip() 
    filtro_rol = request.GET.get('rol')

    if not query and not filtro_rol:
        filtro_rol = 'SUPERVISOR'

    usuarios_list = Usuario.objects.exclude(id=request.user.id).order_by('-date_joined')

    if query:
        usuarios_list = usuarios_list.filter(
            Q(nombres__icontains=query) | 
            Q(apellidos__icontains=query) | 
            Q(dni__icontains=query) |
            Q(email__icontains=query) |
            # Búsqueda Específica de Supervisor:
            Q(profesion__icontains=query) |
            Q(numero_colegiatura__icontains=query) |
            # Búsqueda Específica de Encuestador:
            Q(institucion_procedencia__icontains=query) |
            Q(especialidad__icontains=query)
        ) 

    if filtro_rol:
        usuarios_list = usuarios_list.filter(rol=filtro_rol)

    # 6. Paginación (10 resultados por página)
    paginator = Paginator(usuarios_list, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    base_counts = Usuario.objects.exclude(id=request.user.id)
    total_admins = base_counts.filter(rol='ADMIN').count()
    total_supervisores = base_counts.filter(rol='SUPERVISOR').count()
    total_encuestadores = base_counts.filter(rol='ENCUESTADOR').count()

    context = {
        'usuarios': page_obj,
        'query': query,
        'filtro_rol': filtro_rol,
        'total_admins': total_admins,
        'total_supervisores': total_supervisores,
        'total_encuestadores': total_encuestadores
    }

    return render(request, 'usuarios/lista_usuarios.html', context)
# 2. CREAR Y EDITAR (Vista Híbrida)
@user_passes_test(es_admin)

@user_passes_test(es_admin)
def gestionar_usuario(request, pk=None):
    if pk:
        usuario = get_object_or_404(Usuario, pk=pk)
        titulo = "Editar Usuario"
    else:
        usuario = None
        titulo = "Crear Nuevo Usuario"

    if request.method == 'POST':
        form = UsuarioAdminForm(request.POST, instance=usuario)
        
        if form.is_valid():
            try:
                form.save() # Intentamos guardar
                accion = "actualizado" if pk else "creado"
                messages.success(request, f'Usuario {accion} correctamente.')
                return redirect('lista_usuarios')
            except Exception as e:
                # Si falla al guardar en base de datos (error raro)
                print("❌ ERROR AL GUARDAR EN BD:", e)
                messages.error(request, f'Error de sistema: {e}')
        else:
            # 🚨 AQUÍ ESTÁ LA CLAVE: Imprimimos por qué falló la validación
            print("⚠️ EL FORMULARIO NO ES VÁLIDO. ERRORES DETECTADOS:")
            print(form.errors) 
            # ---------------------------------------------------------
            
            if hasattr(form, 'estilar_errores'): form.estilar_errores()
            messages.error(request, 'Hay errores en el formulario. Revisa los campos rojos.')
    else:
        form = UsuarioAdminForm(instance=usuario)

    return render(request, 'usuarios/form_usuario.html', {'form': form, 'titulo': titulo})


@user_passes_test(es_admin)
def eliminar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    # Evitar que el admin se borre a sí mismo
    if usuario == request.user:
        messages.error(request, "No puedes eliminar tu propia cuenta.")
    else:
        # usuario.is_active = False
        # usuario.save() 
        usuario.delete()
        messages.success(request, "Usuario eliminado correctamente.")
    
    return redirect('lista_usuarios')

# 4. DETALLE DE USUARIO
@user_passes_test(es_admin)
def detalle_usuario(request, pk):
    # Importaciones locales necesarias
    from datetime import timedelta
    from apps.fichas.models import FichaEvaluacion

    usuario = get_object_or_404(Usuario, pk=pk)
    
    # Contexto base
    context = {
        'usuario': usuario,
        'encuestadores_asignados': [],
        'fichas_realizadas': [],
        # Inicializamos contadores en 0 por defecto
        'total_fichas': 0, 'bajo': 0, 'moderado': 0, 'severo': 0, 'critico': 0,
        'filtro_dni': '', 'filtro_fecha': ''
    }

    # =========================================================
    # CASO 1: ES SUPERVISOR (Ver su equipo)
    # =========================================================
    if usuario.rol == 'SUPERVISOR':
        context['encuestadores_asignados'] = usuario.encuestadores_a_cargo.all().order_by('apellidos')

    # =========================================================
    # CASO 2: ES ENCUESTADOR (Ver sus fichas + KPIs + Filtros)
    # =========================================================
    elif usuario.rol == 'ENCUESTADOR':
        # 1. QuerySet Base
        fichas_queryset = FichaEvaluacion.objects.filter(usuario_registra=usuario)

        # 2. Capturar Filtros (Igual que en Supervisor)
        filtro_dni = request.GET.get('dni', '').strip()
        filtro_fecha = request.GET.get('rango_fecha', '')

        # Filtro DNI
        if filtro_dni:
            fichas_queryset = fichas_queryset.filter(dni_evaluado__icontains=filtro_dni)

        # Filtro Fecha
        if filtro_fecha:
            hoy = timezone.now().date()
            if filtro_fecha == 'hoy':
                fichas_queryset = fichas_queryset.filter(fecha_registro__date=hoy)
            elif filtro_fecha == '7dias':
                inicio = hoy - timedelta(days=7)
                fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)
            elif filtro_fecha == 'mes':
                inicio = hoy - timedelta(days=30)
                fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)

        # 3. Calcular KPIs (Sobre la data filtrada)
        context['total_fichas'] = fichas_queryset.count()
        context['bajo'] = fichas_queryset.filter(nivel_riesgo='RIESGO BAJO').count()
        context['moderado'] = fichas_queryset.filter(nivel_riesgo='RIESGO MODERADO').count()
        context['severo'] = fichas_queryset.filter(nivel_riesgo='RIESGO SEVERO').count()
        context['critico'] = fichas_queryset.filter(nivel_riesgo='RIESGO CRÍTICO').count()

        # 4. Obtener datos para la tabla
        if filtro_dni or filtro_fecha:
            context['fichas_realizadas'] = fichas_queryset.order_by('-fecha_registro')
        else:
            context['fichas_realizadas'] = fichas_queryset.order_by('-fecha_registro')[:20]
        
        # Pasar filtros al template
        context['filtro_dni'] = filtro_dni
        context['filtro_fecha'] = filtro_fecha

    return render(request, 'usuarios/detalle_usuario.html', context)

# ==============================================================================
# ZONA SUPERVISOR (Gestión de Equipo)
# ==============================================================================

@user_passes_test(es_supervisor)
def listar_mi_equipo(request):
    query = request.GET.get('q', '').strip()
    
    # 1. Filtro Base: Solo encuestadores ASIGNADOS a mí
    # Si quieres que vea TODOS los encuestadores, quita "supervisor_asignado=request.user"
    equipo = Usuario.objects.filter(rol='ENCUESTADOR', supervisor_asignado=request.user).order_by('-date_joined')

    # 2. Búsqueda local (solo en su equipo)
    if query:
        equipo = equipo.filter(
            Q(nombres__icontains=query) | 
            Q(apellidos__icontains=query) | 
            Q(dni__icontains=query) |
            Q(codigo__icontains=query) |
            Q(institucion_procedencia__icontains=query)
        )

    # 3. Paginación
    paginator = Paginator(equipo, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'equipo': page_obj,
        'query': query,
        'total_miembros': equipo.count()
    }
    # OJO: Creamos una carpeta nueva en templates
    return render(request, 'usuarios/supervisor/lista_equipo.html', context)

# apps/usuarios/views.py

@user_passes_test(es_supervisor)
def crear_encuestador_supervisor(request):
    if request.method == 'POST':
        # Pasamos user_sesion para seguridad
        form = EncuestadorForm(request.POST, user_sesion=request.user)
        if form.is_valid():
            encuestador = form.save(commit=False)
            encuestador.rol = 'ENCUESTADOR'
            encuestador.supervisor_asignado = request.user
            encuestador.save()
            messages.success(request, 'Encuestador creado correctamente.')
            return redirect('listar_mi_equipo')
    else:
        form = EncuestadorForm(user_sesion=request.user)

    # AQUI EL CAMBIO: Apuntamos al nuevo HTML independiente
    return render(request, 'usuarios/supervisor/form_encuestador.html', {
        'form': form,
        'titulo': 'Nuevo Miembro de Equipo'
    })

@user_passes_test(es_supervisor)
def editar_encuestador_supervisor(request, pk):
    encuestador = get_object_or_404(Usuario, pk=pk, rol='ENCUESTADOR', supervisor_asignado=request.user)
    
    if request.method == 'POST':
        # Pasamos user_sesion para seguridad
        form = EncuestadorForm(request.POST, instance=encuestador, user_sesion=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Datos actualizados.')
            return redirect('listar_mi_equipo')
    else:
        form = EncuestadorForm(instance=encuestador, user_sesion=request.user)

    # AQUI EL CAMBIO: Apuntamos al nuevo HTML independiente
    return render(request, 'usuarios/supervisor/form_encuestador.html', {
        'form': form, 
        'titulo': f'Editar a {encuestador.nombres}'
    })



@user_passes_test(es_supervisor)
def ver_detalle_equipo(request, pk):
    # 1. SEGURIDAD: Obtener el encuestador
    encuestador = get_object_or_404(Usuario, pk=pk, rol='ENCUESTADOR', supervisor_asignado=request.user)

    # 2. QUERYSET BASE: Todas las fichas de este usuario
    fichas_queryset = FichaEvaluacion.objects.filter(usuario_registra=encuestador)

    # --- LÓGICA DE FILTROS ---
    # Capturamos los datos que vienen de la URL (del formulario HTML)
    filtro_dni = request.GET.get('dni', '').strip()
    filtro_fecha = request.GET.get('rango_fecha', '')

    # A. Filtrar por DNI
    if filtro_dni:
        fichas_queryset = fichas_queryset.filter(dni_evaluado__icontains=filtro_dni)

    # B. Filtrar por Fecha
    if filtro_fecha:
        hoy = timezone.now().date()
        if filtro_fecha == 'hoy':
            fichas_queryset = fichas_queryset.filter(fecha_registro__date=hoy)
        elif filtro_fecha == '7dias':
            inicio = hoy - timedelta(days=7)
            fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)
        elif filtro_fecha == 'mes':
            inicio = hoy - timedelta(days=30)
            fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)

    # 3. CÁLCULO DE KPIS (Se actualizan según el filtro aplicado)
    # Si filtras por "hoy", los contadores mostrarán solo los riesgos de hoy.
    total_fichas = fichas_queryset.count()
    bajo = fichas_queryset.filter(nivel_riesgo='RIESGO BAJO').count()
    moderado = fichas_queryset.filter(nivel_riesgo='RIESGO MODERADO').count()
    severo = fichas_queryset.filter(nivel_riesgo='RIESGO SEVERO').count()
    critico = fichas_queryset.filter(nivel_riesgo='RIESGO CRÍTICO').count()

    # 4. LISTADO PARA LA TABLA
    # Si el usuario está filtrando, mostramos TODOS los resultados coincidentes.
    # Si NO está filtrando, mostramos solo los últimos 10 o 20 para no saturar.
    if filtro_dni or filtro_fecha:
        fichas_realizadas = fichas_queryset.order_by('-fecha_registro')
    else:
        fichas_realizadas = fichas_queryset.order_by('-fecha_registro')[:20]

    context = {
        'usuario': encuestador,
        'bajo': bajo,
        'moderado': moderado,
        'severo': severo,
        'critico': critico,
        'fichas_realizadas': fichas_realizadas,
        'total_fichas': total_fichas,
        # Pasamos los filtros al template para mantenerlos escritos en los inputs
        'filtro_dni': filtro_dni,
        'filtro_fecha': filtro_fecha
    }

    return render(request, 'usuarios/supervisor/detalle_encuestador.html', context)

@login_required
def eliminar_encuestador_equipo(request, pk):
    # 1. Validación manual del rol para romper el bucle infinito
    if request.user.rol != 'SUPERVISOR':
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('dashboard_admin')

    # 2. Seguridad: Solo permite obtener el encuestador si es del equipo del supervisor
    encuestador = get_object_or_404(
        Usuario, 
        pk=pk, 
        rol='ENCUESTADOR', 
        supervisor_asignado=request.user
    )
    
    nombre = encuestador.nombres

    encuestador.is_active = False  # Marcamos como inactivo en lugar de eliminar
    encuestador.save()
    
    messages.success(request, f'{nombre} ha sido eliminado de tu equipo.')
    return redirect('listar_mi_equipo')


@login_required
def exportar_excel_supervisor(request):
    # 1. BASE: Obtener todas las fichas inicialmente
    # (Si quieres restringir que un encuestador solo baje las suyas, filtra aquí por request.user)
    if request.user.rol == 'ENCUESTADOR':
        fichas_queryset = FichaEvaluacion.objects.filter(usuario_registra=request.user)
    else:
        # Admin o Supervisor inician viendo todo
        fichas_queryset = FichaEvaluacion.objects.all().select_related('usuario_registra', 'institucion')

    # ======================================================
    # 2. APLICACIÓN DE FILTROS (Misma lógica que tu vista HTML)
    # ======================================================
    
    # A. Filtro por USUARIO ESPECÍFICO (Cuando el supervisor entra al detalle de alguien)
    usuario_id = request.GET.get('usuario_id')
    if usuario_id:
        fichas_queryset = fichas_queryset.filter(usuario_registra_id=usuario_id)

    # B. Filtro por DNI (Buscador)
    filtro_dni = request.GET.get('dni', '').strip()
    if filtro_dni:
        fichas_queryset = fichas_queryset.filter(dni_evaluado__icontains=filtro_dni)

    # C. Filtro por FECHAS
    filtro_fecha = request.GET.get('rango_fecha', '')
    if filtro_fecha:
        hoy = timezone.now().date()
        if filtro_fecha == 'hoy':
            fichas_queryset = fichas_queryset.filter(fecha_registro__date=hoy)
        elif filtro_fecha == '7dias':
            inicio = hoy - timedelta(days=7)
            fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)
        elif filtro_fecha == 'mes':
            inicio = hoy - timedelta(days=30)
            fichas_queryset = fichas_queryset.filter(fecha_registro__date__gte=inicio)

    # Ordenamos por fecha descendente
    fichas_queryset = fichas_queryset.order_by('-fecha_registro')

    # ======================================================
    # 3. CREACIÓN DEL ARCHIVO EXCEL
    # ======================================================
    
    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Fichas"

    # Definir Cabeceras
    headers = [
        "Fecha Registro",
        "Encuestador",
        "Evaluado (Estudiante)",
        "DNI",
        "Edad",
        "Institución",
        "Puntaje Total",
        "Nivel de Riesgo"
    ]

    # Escribir Cabeceras y Estilizar
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Azul tipo Tailwind

    for cell in ws[1]: # Primera fila
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Escribir Datos
    for ficha in fichas_queryset:
        # Formatear fecha para que no salga con microsegundos
        fecha_str = ficha.fecha_registro.strftime('%d/%m/%Y %H:%M') if ficha.fecha_registro else '-'
        
        # Concatenar nombres
        nombre_completo = f"{ficha.nombres_evaluado} {ficha.apellidos_evaluado}"
        encuestador_nombre = f"{ficha.usuario_registra.nombres} {ficha.usuario_registra.apellidos}" if ficha.usuario_registra else "Desconocido"
        colegio = ficha.institucion.nombre if ficha.institucion else "Sin Institución"

        row = [
            fecha_str,              # Col A
            encuestador_nombre,     # Col B
            nombre_completo,        # Col C
            ficha.dni_evaluado,     # Col D
            ficha.edad_evaluado,    # Col E
            colegio,                # Col F
            ficha.puntaje_total,    # Col G
            ficha.nivel_riesgo      # Col H
        ]
        ws.append(row)

    # Ajustar ancho de columnas automáticamente (Opcional pero recomendado)
    column_widths = [20, 30, 35, 15, 10, 30, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ======================================================
    # 4. RESPUESTA HTTP (Descarga)
    # ======================================================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Riesgo_Social_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# Redireccionamiento 


def error_404_view(request, exception):
    """Si la URL no existe, lo mandamos al home dinámico."""
    if request.user.is_authenticated:
        return redirect('home') # Usa el 'name' definido en tu urls.py
    return redirect('login')

def error_403_view(request, exception=None):
    """Si no tiene permiso para una URL existente, lo regresamos a su dashboard."""
    if request.user.is_authenticated:
        return redirect('home')
    return redirect('login')